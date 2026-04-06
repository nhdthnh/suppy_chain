import openpyxl
import os

def check():
    path = r"C:\Users\OQR\Desktop\supply_chain\templates\Pre-Order\LIEBU\0403_2026_OQR_LIEBU copy.xlsx"
    if not os.path.exists(path):
        print(f"Path not found: {path}")
        return
    
    try:
        # read_only=True is much faster for large files
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "master data" not in wb.sheetnames:
            print("Sheet 'master data' not found.")
            return
            
        ws = wb["master data"]
        print("\n--- MASTER DATA PREVIEW (First 15 rows) ---")
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=6)):
            vals = [str(c.value) for c in row]
            print(f"Row {idx+1}: {vals}")
        wb.close()
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    check()
