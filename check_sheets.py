import openpyxl, os

path = r"C:\Users\OQR\Desktop\supply_chain\templates\Pre-Order\LIEBU\0403_2026_OQR_LIEBU copy.xlsx"
wb = openpyxl.load_workbook(path)  # NOT read_only, to see hidden sheets
print("All sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    vis = ws.sheet_state  # 'visible', 'hidden', 'veryHidden'
    print(f"  '{name}' -> state={vis}, max_row={ws.max_row}, max_col={ws.max_column}")
    if "master" in name.lower():
        print(f"  --- MASTER DATA FOUND: '{name}' ---")
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, max_col=6)):
            vals = [repr(c.value) for c in row]
            print(f"  Row {idx+1}: {vals}")
wb.close()
