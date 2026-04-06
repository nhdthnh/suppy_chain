"""
pages/view_templates.py — Xem trước các file Excel Templates.

Flow:
  1. Chọn loại biểu mẫu (thư mục con trong templates/)
  2. Chọn vendor
  3. Chọn file .xlsx
  4. Render bảng HTML preview (hỗ trợ merged cells, colors, fonts)
"""

import os

import streamlit as st
from utils.styles import page_header, section_label, divider


# ─────────────────────────────────────────────────────────────
# EXCEL → HTML RENDERER
# ─────────────────────────────────────────────────────────────

def _excel_to_html(filepath: str, sheet_name: str) -> str:
    """Render Excel sheet thành bảng HTML (fallback khi không có xlsx2html).

    Hỗ trợ:
      · Merged cells (rowspan/colspan)
      · Cell fill color
      · Font bold/italic/color
      · Text alignment
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import html as html_lib

    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        return "<em>Sheet không tồn tại.</em>"
    ws = wb[sheet_name]

    # ── Xử lý merged cells ───────────────────────────────
    merged = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r == mr.min_row and c == mr.min_col:
                    merged[(r, c)] = {
                        "rowspan": mr.max_row - mr.min_row + 1,
                        "colspan": mr.max_col - mr.min_col + 1,
                    }
                else:
                    merged[(r, c)] = "slave"

    # ── Style cho index cells ────────────────────────────
    IDX_STYLE = (
        "background:#f4f4f5;text-align:center;font-weight:600;color:#555;"
        "border:1px solid #d1d5db;padding:3px 5px;"
        "font-family:'JetBrains Mono',monospace;font-size:0.65rem;"
    )

    # ── Build HTML table ─────────────────────────────────
    html = [
        '<table style="border-collapse:collapse;'
        "font-family:'Source Sans',sans-serif;font-size:0.78rem;"
        'border:1px solid #e2e8f0;color:#222;background:#fff;'
        'width:max-content;">',
    ]

    # Header row (column letters)
    html.append("<tr>")
    html.append(
        f'<td style="{IDX_STYLE}border-bottom:2px solid #ccc;'
        f'width:35px;"></td>'
    )
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        col_dim = ws.column_dimensions.get(letter)
        width = col_dim.width if col_dim and col_dim.width else 8.43
        w_px = int(width * 8.5)
        html.append(
            f'<td style="{IDX_STYLE}border-bottom:2px solid #ccc;'
            f'width:{w_px}px;min-width:{w_px}px;">{letter}</td>'
        )
    html.append("</tr>")

    # Data rows
    for r in range(1, ws.max_row + 1):
        html.append('<tr style="height:24px;">')
        html.append(
            f'<td style="{IDX_STYLE}border-right:2px solid #ccc;">'
            f"{r}</td>"
        )
        for c in range(1, ws.max_column + 1):
            status = merged.get((r, c))
            if status == "slave":
                continue

            cell = ws.cell(r, c)
            val = cell.value if cell.value is not None else ""
            val = html_lib.escape(str(val)) if val != "" else "&nbsp;"

            rowspan = colspan = 1
            if isinstance(status, dict):
                rowspan = status["rowspan"]
                colspan = status["colspan"]

            styles = ["border:1px solid #e5e7eb;padding:4px 7px;"]

            # Cell fill color
            if cell.fill and cell.fill.start_color:
                sc = cell.fill.start_color
                rgb = getattr(sc, "rgb", None)
                if rgb and isinstance(rgb, str) and rgb != "00000000":
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if rgb.upper() not in ("FFFFFF", "000000"):
                        styles.append(f"background-color:#{rgb};")

            # Font styles
            if cell.font:
                if cell.font.bold:
                    styles.append("font-weight:bold;")
                if cell.font.italic:
                    styles.append("font-style:italic;")
                fc = cell.font.color
                frgb = fc.rgb if fc and hasattr(fc, "rgb") else None
                if frgb and isinstance(frgb, str) and frgb != "00000000":
                    if len(frgb) == 8:
                        frgb = frgb[2:]
                    if frgb.upper() != "000000":
                        styles.append(f"color:#{frgb};")

            # Alignment
            if cell.alignment:
                if cell.alignment.horizontal:
                    h = cell.alignment.horizontal.replace("centre", "center")
                    styles.append(f"text-align:{h};")
                if cell.alignment.vertical:
                    v = cell.alignment.vertical.replace("centre", "center")
                    styles.append(f"vertical-align:{v};")

            r_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ""
            c_attr = f' colspan="{colspan}"' if colspan > 1 else ""
            html.append(
                f'<td{r_attr}{c_attr} style="{"".join(styles)}">{val}</td>'
            )
        html.append("</tr>")

    html.append("</table>")
    return "".join(html)


# ─────────────────────────────────────────────────────────────
# RENDER
# ─────────────────────────────────────────────────────────────

def render() -> None:
    """Render trang xem Templates Excel."""
    st.markdown(
        page_header("supply_chain / templates", "XEM TEMPLATES EXCEL"),
        unsafe_allow_html=True,
    )

    base = os.path.join(os.path.dirname(__file__), "..", "templates")
    if not os.path.exists(base):
        st.error("Thư mục `templates` không tồn tại.")
        return

    types = [
        d for d in os.listdir(base)
        if os.path.isdir(os.path.join(base, d)) and not d.startswith(".")
    ]
    if not types:
        st.warning("Không tìm thấy thư mục con nào trong `templates`.")
        return

    # ── Chọn loại + vendor + file ─────────────────────────
    c1, c2 = st.columns([3, 7])
    with c1:
        selected_type = st.selectbox(
            "📂 Loại biểu mẫu", options=types, key="tmpl_type_select"
        )

    type_path = os.path.join(base, selected_type)
    vendors = [
        d for d in os.listdir(type_path)
        if os.path.isdir(os.path.join(type_path, d)) and not d.startswith(".")
    ]
    template_file = None

    with c2:
        if not vendors:
            st.info(f"Chưa có thư mục Vendor cho `{selected_type}`.")
        else:
            selected_vendor = st.selectbox(
                "🏢 Vendor", options=vendors, key="tmpl_vendor_select"
            )
            vendor_path = os.path.join(type_path, selected_vendor)
            files = [f for f in os.listdir(vendor_path) if f.endswith(".xlsx")]
            if not files:
                st.warning(
                    f"Chưa có file `.xlsx` trong "
                    f"`templates/{selected_type}/{selected_vendor}`."
                )
            else:
                selected_file = st.selectbox(
                    "📄 File Template", options=files, key="tmpl_file_select"
                )
                template_file = os.path.join(vendor_path, selected_file)

    st.markdown(divider(), unsafe_allow_html=True)

    if not template_file:
        return

    # ── Render preview ────────────────────────────────────
    st.markdown(
        section_label(
            "📊", f"BẢN XEM TRƯỚC: {os.path.basename(template_file)}"
        ),
        unsafe_allow_html=True,
    )

    try:
        from openpyxl import load_workbook

        wb_sheets = load_workbook(
            template_file, read_only=True
        ).sheetnames
        if not wb_sheets:
            st.warning("File Excel này không có sheet nào.")
            return

        sc1, _ = st.columns([3, 7])
        with sc1:
            selected_sheet = st.selectbox(
                "🔍 Sheet", options=wb_sheets, key="tmpl_sheet_select"
            )

        with st.spinner("Đang kết xuất biểu mẫu…"):
            try:
                import io
                from xlsx2html import xlsx2html

                html_buf = io.StringIO()
                xlsx2html(template_file, html_buf, sheet=selected_sheet)
                html_view = html_buf.getvalue()
                style_override = (
                    "<style>"
                    "table{border-collapse:collapse!important;"
                    "width:max-content!important;"
                    "font-family:'Source Sans',sans-serif!important;"
                    "font-size:0.78rem!important;}"
                    "td{border:1px solid #e5e7eb!important;"
                    "padding:4px 8px!important;}"
                    "</style>"
                )
                html_view = style_override + html_view
            except ImportError:
                html_view = _excel_to_html(template_file, selected_sheet)

        st.markdown(
            f'<div style="overflow-x:auto;overflow-y:auto;max-height:620px;'
            f"border:1px solid rgba(128,128,128,0.18);border-radius:7px;"
            f'background:#fff;padding:10px;">'
            f"{html_view}</div>",
            unsafe_allow_html=True,
        )

    except Exception as e:
        st.error(f"Lỗi đọc file Excel: {e}")
