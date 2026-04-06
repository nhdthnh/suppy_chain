"""
modules/export_pre_order.py — Điền sản phẩm vào template Excel (PO + CI).

Cách hoạt động:
  1. Load config từ config/export_pre_order.json (fallback default)
  2. Pass 1 (structural): insert thêm rows nếu cần
  3. Pass 2 (data): ghi sản phẩm, formulas, static fields
  4. Pass 3 (sync): đọc giá từ 'master data' (kể cả hidden) để sync DB
  5. Trả về (bytes, sync_data)

Config format — xem config/export_pre_order.json
"""

import copy
import io
import json
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import date

import pandas as pd
from openpyxl import load_workbook

_HERE = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_HERE, "..", "config", "export_pre_order.json")


# ─────────────────────────────────────────────────────────────
# DEFAULT CONFIG (fallback khi file JSON bị xóa/lỗi)
# ─────────────────────────────────────────────────────────────

_DEFAULT_CONFIG = {
    "sheets": [
        {
            "name": "PO",
            "first_row": 15,
            "ref_row": 15,
            "fields": {
                "F2": {"type": "date", "format": "%d/%m/%Y"},
                "F3": {"type": "po_number"},
            },
            "columns": [
                {"col": 1, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$B, 2, 0)"},
                {"col": 2, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$C, 3, 0)"},
                {"col": 3, "type": "field", "value": "description"},
                {"col": 4, "type": "field", "value": "quantity", "cast": "int"},
                {"col": 5, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$D, 4, 0)"},
                {"col": 6, "type": "formula", "value": "=D{row}*E{row}"},
                {"col": 7, "type": "field", "value": "note"},
            ],
            "totals": [
                {"col": 4, "formula": "=SUM(D{start}:D{end})"},
                {"col": 6, "formula": "=SUM(F{start}:F{end})"},
            ],
            "total_search_text": "TOTAL",
        },
        {
            "name": "CI",
            "first_row": 12,
            "ref_row": 12,
            "fields": {
                "E4": {"type": "date", "format": "%d/%m/%Y"},
            },
            "columns": [
                {"col": 1, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$B, 2, 0)"},
                {"col": 2, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$C, 3, 0)"},
                {"col": 3, "type": "field", "value": "description"},
                {"col": 4, "type": "field", "value": "quantity", "cast": "int"},
                {"col": 5, "type": "formula",
                 "value": "=VLOOKUP(C{row}, 'master data'!$A:$D, 4, 0)"},
                {"col": 6, "type": "formula", "value": "=E{row}*D{row}"},
                {"col": 7, "type": "formula",
                 "value": "=F{row}*{deposit_pct}%"},
            ],
            "totals": [
                {"col": 4, "formula": "=SUM(D{start}:D{end})"},
                {"col": 6, "formula": "=SUM(F{start}:F{end})"},
                {"col": 7, "formula": "=SUM(G{start}:G{end})"},
            ],
            "total_search_text": "TOTAL",
        },
    ]
}


def _load_config() -> dict:
    """Load config từ file JSON, fallback sang default."""
    if os.path.exists(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return _DEFAULT_CONFIG


# ─────────────────────────────────────────────────────────────
# CELL / ROW HELPERS
# ─────────────────────────────────────────────────────────────

def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)


def _copy_row_style(ws_src, src_row, ws_dst, dst_row, max_col=10):
    for col in range(1, max_col + 1):
        _copy_cell_style(ws_src.cell(src_row, col), ws_dst.cell(dst_row, col))


def _safe_set(ws, row, col, value):
    c = ws.cell(row, col)
    if type(c).__name__ != "MergedCell":
        c.value = value


def _find_total_row(ws, first_row, search_text="TOTAL"):
    for r in range(first_row, ws.max_row + 5):
        if ws.cell(r, 1).value == search_text:
            return r
    return first_row + 4


def _hide_row(ws, row):
    ws.row_dimensions[row].height = 0
    ws.row_dimensions[row].hidden = True
    for col in range(1, 10):
        _safe_set(ws, row, col, None)


def _show_row(ws, row, height):
    ws.row_dimensions[row].height = height
    ws.row_dimensions[row].hidden = False


# ─────────────────────────────────────────────────────────────
# PASS 1: STRUCTURAL — INSERT ROWS NẾU CẦN
# ─────────────────────────────────────────────────────────────

def _structural_pass(tmpl: bytes, n: int, config: dict) -> bytes:
    wb = load_workbook(io.BytesIO(tmpl))
    for scfg in config.get("sheets", []):
        sname = scfg["name"]
        first = scfg["first_row"]
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        total_row = _find_total_row(ws, first, scfg.get("total_search_text", "TOTAL"))
        diff = n - (total_row - first)
        if diff > 0:
            for mr in [str(m) for m in ws.merged_cells.ranges if m.min_row >= first]:
                ws.merged_cells.remove(mr)
            ws.insert_rows(total_row, diff)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# PASS 2: WRITE DATA
# ─────────────────────────────────────────────────────────────

def _write_sheet(ws, ws_ref, items, po_date, po_number, deposit_pct, cfg):
    first = cfg.get("first_row", 15)
    n = len(items)
    total_text = cfg.get("total_search_text", "TOTAL")
    total_row = _find_total_row(ws, first, total_text)

    # Static fields
    for cell_addr, fcfg in cfg.get("fields", {}).items():
        ftype = fcfg.get("type")
        val = None
        if ftype == "date":
            val = po_date.strftime(fcfg.get("format", "%d/%m/%Y"))
        elif ftype == "po_number":
            val = po_number
        if val is not None:
            ws[cell_addr] = val

    # Clear data area
    for r in range(first, total_row):
        for col in range(1, 10):
            _safe_set(ws, r, col, None)

    # Write items
    ref_row = cfg.get("ref_row", first)
    data_h = ws_ref.row_dimensions[ref_row].height or 47.25
    pct = int(deposit_pct) if deposit_pct == int(deposit_pct) else deposit_pct

    for idx, item in enumerate(items):
        r = first + idx
        _show_row(ws, r, data_h)
        _copy_row_style(ws_ref, ref_row, ws, r)
        for col_cfg in cfg.get("columns", []):
            c_idx = col_cfg["col"]
            ctype = col_cfg["type"]
            val_template = col_cfg.get("value", "")
            if ctype == "formula":
                try:
                    val = val_template.format(row=r, deposit_pct=pct)
                except Exception:
                    val = val_template
                ws.cell(r, c_idx).value = val
            elif ctype == "field":
                val = item.get(val_template, "")
                if col_cfg.get("cast") == "int":
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        val = 0
                ws.cell(r, c_idx).value = val

    # Hide extra rows
    for r in range(first + n, total_row):
        _hide_row(ws, r)

    # Update TOTAL formulas
    for tcfg in cfg.get("totals", []):
        c_idx = tcfg["col"]
        try:
            val = tcfg["formula"].format(start=first, end=first + n - 1)
        except Exception:
            val = tcfg["formula"]
        ws.cell(total_row, c_idx).value = val


# ─────────────────────────────────────────────────────────────
# PRICE LOOKUP (đọc từ sheet ẩn 'master data')
# ─────────────────────────────────────────────────────────────

def normalize_key(s: str) -> str:
    """Chuẩn hóa key: lowercase, bỏ ngoặc đơn, bỏ ký tự đặc biệt."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s*\(.*?\)\s*$", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def _find_master_sheet_name(file_path: str) -> str | None:
    """Tìm tên chính xác của sheet chứa 'master' (kể cả hidden) qua XML."""
    try:
        with zipfile.ZipFile(file_path, "r") as z:
            with z.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for s in root.findall(".//m:sheet", ns):
                    name = s.get("name", "")
                    if "master" in name.lower():
                        return name
    except Exception:
        pass
    return None


def get_master_prices_from_file(file_path: str) -> dict:
    """Đọc giá từ sheet 'master data' (kể cả bị ẩn) bằng pandas.
    Returns: dict {normalized_key: price}
    """
    sheet_name = _find_master_sheet_name(file_path)
    if not sheet_name:
        return {}
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        if df.empty:
            return {}
        price_map = {}
        # Cột A (idx 0) = Description/Key,  Cột D (idx 3) = Price
        for _, row in df.iterrows():
            key_raw = row.iloc[0]
            price_raw = row.iloc[3] if len(row) > 3 else None
            if key_raw is not None and price_raw is not None:
                try:
                    price = float(price_raw)
                except (ValueError, TypeError):
                    continue
                norm = normalize_key(str(key_raw))
                if norm:
                    price_map[norm] = price
                raw = str(key_raw).strip().lower()
                if raw and raw not in price_map:
                    price_map[raw] = price
        return price_map
    except Exception:
        return {}


def get_master_prices(template_bytes: bytes) -> dict:
    """Wrapper: bytes → temp file → get_master_prices_from_file."""
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(template_bytes)
            tmp_path = tmp.name
        result = get_master_prices_from_file(tmp_path)
        os.remove(tmp_path)
        return result
    except Exception:
        return {}


# ─────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────

def build_pre_order_bytes(
    template_bytes: bytes,
    items: list[dict],
    po_date: date,
    vendor: dict,
    deposit_pct: float = 70.0,
) -> tuple[bytes, list[dict]]:
    """Tạo file Excel Pre-Order và trả về dữ liệu đã tính toán.

    Returns:
        (excel_bytes, sync_data)
        sync_data: [{Barcode, Description, Qty, Unit Price, Amount, Deposit}]
    """
    if not items:
        return template_bytes, []

    dd = po_date.strftime("%d")
    mm = po_date.strftime("%m")
    yyyy = po_date.strftime("%Y")
    short = (vendor.get("short_name") or vendor.get("name", "VENDOR")).strip()
    po_number = f"{dd}{mm}_{yyyy}_OQR_{short}"

    config = _load_config()
    adjusted = _structural_pass(template_bytes, len(items), config)

    wb = load_workbook(io.BytesIO(adjusted))
    wb_ref = load_workbook(io.BytesIO(template_bytes))

    # Ghi dữ liệu vào các sheet (PO, CI)
    for scfg in config.get("sheets", []):
        name = scfg["name"]
        if name in wb.sheetnames:
            _write_sheet(
                wb[name], wb_ref[name],
                items, po_date, po_number, deposit_pct, scfg,
            )

    # Tính toán sync data từ master data (kể cả hidden sheet)
    price_map = get_master_prices(template_bytes)
    sync_data = []
    for it in items:
        desc_raw = it.get("description", "")
        bc_raw = it.get("barcode", "")

        # Smart match: normalized desc → barcode → raw desc
        price = price_map.get(normalize_key(desc_raw), 0)
        if price == 0:
            price = price_map.get(str(bc_raw).strip().lower(), 0)
        if price == 0:
            price = price_map.get(str(desc_raw).strip().lower(), 0)

        qty = int(it.get("quantity", 1))
        amount = qty * price
        deposit = amount * deposit_pct / 100

        sync_data.append({
            "Barcode": bc_raw,
            "Description": desc_raw,
            "Qty": qty,
            "Unit Price": price,
            "Amount": amount,
            "Deposit": deposit,
        })

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), sync_data
